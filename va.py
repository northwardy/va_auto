import xml.etree.ElementTree as etree
import sys
import time
import xlsxwriter
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
import xlrd
import csv
import shutil
import os

def merged():
	first = 1
	for fileName in os.listdir("."):
		if ".nessus" in fileName:
			print(":: Parsing", fileName)
			if first:
				mainTree = etree.parse(fileName)
				report = mainTree.find('Report')
				report.attrib['name'] = 'Merged Report'
				first = 0
			else:
				tree = etree.parse(fileName)
				for host in tree.findall('.//ReportHost'):
					existing_host = report.find(".//ReportHost[@name='"+host.attrib['name']+"']")
					if not existing_host:
						print "adding host: " + host.attrib['name']
						report.append(host)
					else:
						for item in host.findall('ReportItem'):
							if not existing_host.find("ReportItem[@port='"+ item.attrib['port'] +"'][@pluginID='"+ item.attrib['pluginID'] +"']"):
								print "adding finding: " + item.attrib['port'] + ":" + item.attrib['pluginID']
                                                                existing_host.append(item)
	                print(":: => done.")

	if "nss_report" in os.listdir("."):
		shutil.rmtree("nss_report")

	os.mkdir("nss_report")
	mainTree.write("nss_report/report.nessus", encoding="utf-8", xml_declaration=True)

def va():
	wb = Workbook()
	ws = wb.active

	ws['A1'] = "_time"
	ws['B1'] = "Host"
	ws['C1'] = "Port"
	ws['D1'] = "Protocol"
	ws['E1'] = "Risk"
	ws['F1'] = "PID"
	ws['G1'] = "Vulnerability"
	ws['H1'] = "Description"
	ws['I1'] = "Solution"
	ws['J1'] = "CVE"
	ws['K1'] = "Vulnerability_publication"
	ws['L1'] = "Exploit_available"
	ws['M1'] = "Patch_publication"
	ws['N1'] = "CVSS"
	ws['O1'] = "SO"
	ws['P1'] = "Platform"
	ws['Q1'] = "hostname"

	xmlFile = "nss_report/report.nessus" 

	tree = etree.parse(xmlFile)
	present = datetime.now()
	delay = datetime.now() - timedelta(days=90)

	root = tree.getroot()
	so = ""
	platform = ""
	host_fqdn = ""
	host_ip = ""

	A = 'A'
	B = 'B'
	C = 'C'
	D = 'D'
	E = 'E'
	F = 'F'
	G = 'G'
	H = 'H'
	I = 'I'
	J = 'J'
	K = 'K'
	L = 'L'
	M = 'M'
	N = 'N'
	O = 'O'
	P = 'P'
	Q = 'Q'
        R = 'R'
        S = 'S'
        linha = 2


	print "Processando: " +xmlFile

	for host in root.iter('ReportHost'):
		address = host.attrib['name']
		#print host._children[000]._children[11].text

		# Outra possibilidade para obter o
		for report in host.findall('HostProperties'):
			for x in report:
				#print x.attrib['name']
				if x.attrib['name'] == 'operating-system':
					so = x.text
					break

		for report in host.findall('HostProperties'):
			for y in report:
				#print x.attrib['name']
				if y.attrib['name'] == 'os':
					platform = y.text
					break

		for report in host.findall('HostProperties'):
			for z in report:
				#print x.attrib['name']
				if z.attrib['name'] == 'host-ip':
					host_ip = z.text
					break

	        if address is not None:
	        	for vuln in host.findall('ReportItem'):
			        port = str(vuln.attrib['port'])
			        protocol = vuln.attrib['protocol']
			        risk_factor = vuln.find('risk_factor').text
			        pid = vuln.attrib['pluginID']
			        name = vuln.attrib['pluginName']
			        description = vuln.find('description').text

			        colunaA = A+str(linha)
			        colunaB = B+str(linha)
			        colunaC = C+str(linha)
			        colunaD = D+str(linha)
			        colunaE = E+str(linha)
			        colunaF = F+str(linha)
			        colunaG = G+str(linha)
		        	colunaH = H+str(linha)
			        colunaI = I+str(linha)
			        colunaJ = J+str(linha)
			        colunaK = K+str(linha)
			        colunaL = L+str(linha)
			        colunaM = M+str(linha)
			        colunaN = N+str(linha)
			        colunaO = O+str(linha)
			        colunaP = P+str(linha)
			        colunaQ = Q+str(linha)
                                colunaR = R+str(linha)
                                colunaS = S+str(linha)

			        cve = vuln.find('cve')

			        if cve is not None:
				        each_line =  cve.text
				        cvestr = each_line.replace("\n", ",").strip()
			        else:
				        cvestr = "N/A"

                                cvssv2 = vuln.find('cvss_vector')

                                if cvssv2 is not None:
                                        each_line = cvssv2.text
                                        cvssv2str = each_line.replace("\n", ",").strip()
                                else:
                                        cvssv2str = "N/A"

                                cvssv3 = vuln.find('cvss3_vector')

                                if cvssv3 is not None:
                                        each_line = cvssv3.text
                                        cvssv3str = each_line.replace("\n", ",").strip()
                                else:
                                        cvssv3str = "N/A"

			        solution = vuln.find('solution')

			        if solution is not None:
				        each_line =  solution.text
				        solution = each_line.replace("\n", ",").strip()
			        else:
				        solution = "N/A"

			        if so is None:
				        so = "N/A"

		        	if platform is None:
			        	so = "N/A"

			        exploit_available = ''
			        exploit_availableFull = vuln.getiterator('exploit_available')
			        for exp in exploit_availableFull:
				        exploit_available = exp.text

			        vuln_publication_date = ''
			        vuln_publication_dateFull = vuln.getiterator('vuln_publication_date')
		        	for vul_date in vuln_publication_dateFull:
			        	vuln_publication_date = vul_date.text

			        patch_publication_date = ''
			        patch_publication_dataFull = vuln.getiterator('patch_publication_date')
			        for patch in patch_publication_dataFull:
				        patch_publication_date  = patch.text

			        cvss_base_score = ''
			        cvss_base_scoreFull = vuln.getiterator('cvss_base_score')
			        for cvss in cvss_base_scoreFull:
				        cvss_base_score  = cvss.text

			        _time = ''
			        _time = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

			        ws[colunaA] = _time
			        ws[colunaB] = host_ip
			        ws[colunaC] = port
			        ws[colunaD] = protocol
		        	ws[colunaE] = risk_factor
		        	ws[colunaF] = pid
		        	ws[colunaG] = name
	        		ws[colunaH] = description
		        	ws[colunaI] = solution
		        	ws[colunaJ] = cvestr
		        	ws[colunaK] = vuln_publication_date
		        	ws[colunaL] = exploit_available
		        	ws[colunaM] = patch_publication_date
		        	ws[colunaN] = cvss_base_score
		        	ws[colunaO] = so
		        	ws[colunaP] = platform
		        	ws[colunaQ] = address
                                ws[colunaR] = cvssv3str
                                ws[colunaS] = cvssv2str

		        	linha += 1
		        	print "[+] ----------------------------------------"

	wb.save("Vulnerabilidades.xlsx")

def csv_convert():
    print "Gerando arquivo csv..."
    wb = xlrd.open_workbook('Vulnerabilidades.xlsx')
    sh = wb.sheet_by_name('Sheet')
    Vulnerabilidades = open('Vulnerabilidades.csv', 'w')
    wr = csv.writer(Vulnerabilidades, quoting=csv.QUOTE_ALL)

    for rownum in range(sh.nrows):
        wr.writerow(sh.row_values(rownum))

    Vulnerabilidades.close()

def remove():
    print "Removendo lixo..."
    os.remove("Vulnerabilidades.xlsx")
    os.remove("nss_report/report.nessus")
    os.rmdir("nss_report")

# run the functions

merged()
va()
csv_convert()
remove ()
